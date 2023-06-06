import openpyxl
from flask import Flask, render_template

workbook = openpyxl.load_workbook('C:/Users/황지혁/Desktop/MRMR/flask_excel.xlsx')
sheet = workbook.active

# 열의 개수와 행의 개수
num_rows = 1000
num_cols = 7

# 딕셔너리
data_dict = {}
accuracyList = []
time_set = set()

# 각 col에 대해 반복(정확도 col 제외한 나머지 col에 대해 가장 최신 정보로 key-value(sec-value) 딕셔너리 완성)
for col in range(2, num_cols+1):
    # 열 이름 가져오기
    col_name = sheet.cell(row=1, column=col).value

    # 딕셔너리 초기화
    column_dict = {}

    # 각 행에 대해 반복
    for row in range(2, num_rows + 1):
        # 시간 정보 가져오기
        time = sheet.cell(row=row, column=1).value

        # 값 가져오기
        value = sheet.cell(row=row, column=col).value

        # 값이 존재하면 딕셔너리에 저장
        if value is not None:
            column_dict = {}
            column_dict[time] = value
            time_set.add(time)

        # 해당 열이 마지막 행에 도달하면 반복 종료
        if row == num_rows:
            break

    # 열에 대한 딕셔너리 저장
    data_dict[col_name] = column_dict

col = num_cols + 1
for row in range(2, num_rows + 1):
    value = sheet.cell(row=row, column=col).value

    if value is not None:
        accuracyList.append(value)

    if row == num_rows:
            break

key_list = []
value_list = []

for dict_key, dict_value in data_dict.items():
    key_list.append(list(dict_value.keys())[0])
    value_list.append(dict_value[list(dict_value.keys())[0]])

# 가장 최신 time
time = key_list[0]

# n순위 차량 정보 리스트 (어떤차인지, cpu)
firstList = [value_list[0], value_list[1]]
secondList = [value_list[2], value_list[3]]
thirdList = [value_list[4], value_list[5]]

# 그래프 x축에 사용될 time_set 리스트로 변환 후 렌더링
time_set = list(time_set)
time_set.sort()


# 1순위 차량에 대해 x,y 좌표 가져오기
coordinate_workbook = openpyxl.load_workbook('C:/Users/황지혁/Documents/카카오톡 받은 파일/4intersection (2)/4intersection/4intersection/4intersection/intersection.xlsx')

# value_list[0]이 'taxi%'로 시작하는 경우 >> car_coordinateX/Y sheet에서 x,y 좌표 추출
if value_list[0].startswith('taxi'):
    # taxi 몇 번 차량인지 number_taxi에 저장
    number_taxi = int(value_list[0][4:])
    taxi_coordinateX = int(coordinate_workbook.worksheets[1].cell(row=time+1, column=number_taxi+1).value)
    taxi_coordinateY = int(coordinate_workbook.worksheets[2].cell(row=time+1, column=number_taxi+1).value)
    firstList.append(taxi_coordinateX)
    firstList.append(taxi_coordinateY)

# value_list[0]이 'bus%'로 시작하는 경우 >> bus_coordinateX/Y sheet에서 x,y 좌표 추출
elif value_list[0].startswith('bus'):
    # bus 몇 번 차량인지 number_bus에 저장
    number_bus = int(value_list[0][3:])
    bus_coordinateX = int(coordinate_workbook.worksheets[5].cell(row=time+1, column=number_bus+1).value)
    bus_coordinateY = int(coordinate_workbook.worksheets[6].cell(row=time+1, column=number_bus+1).value)
    firstList.append(bus_coordinateX)
    firstList.append(bus_coordinateY)



# Flask 앱 생성
app = Flask(__name__)


@app.route('/')
def index():
    return render_template('index.html', data=data_dict, accuracy=accuracyList, time=time, time_set=time_set, firstList=firstList, secondList=secondList, thirdList=thirdList)

@app.route('/help')
def help():
    return render_template('help.html')

@app.route('/update')
def update():
    workbook = openpyxl.load_workbook('C:/Users/황지혁/Desktop/MRMR/flask_excel.xlsx')
    sheet = workbook.active

    num_rows = 1000
    num_cols = 7

    data_dict = {}
    accuracyList = []
    time_set = set()

    for col in range(2, num_cols+1):
        col_name = sheet.cell(row=1, column=col).value
        column_dict = {}

        for row in range(2, num_rows + 1):
            time = sheet.cell(row=row, column=1).value
            value = sheet.cell(row=row, column=col).value

            if value is not None:
                column_dict = {}
                column_dict[time] = value
                time_set.add(time)

            if row == num_rows:
                break

        data_dict[col_name] = column_dict

    col = num_cols + 1
    for row in range(2, num_rows + 1):
        value = sheet.cell(row=row, column=col).value

        if value is not None:
            accuracyList.append(value)

        if row == num_rows:
            break
    key_list = []
    value_list = []

    for dict_key, dict_value in data_dict.items():
        key_list.append(list(dict_value.keys())[0])
        value_list.append(dict_value[list(dict_value.keys())[0]])

    time = key_list[0]
    firstList = [value_list[0], value_list[1]]
    secondList = [value_list[2], value_list[3]]
    thirdList = [value_list[4], value_list[5]]
    time_set = list(time_set)
    time_set.sort()

    coordinate_workbook = openpyxl.load_workbook(
        'C:/Users/황지혁/Documents/카카오톡 받은 파일/4intersection (2)/4intersection/4intersection/4intersection/intersection.xlsx')

    if value_list[0].startswith('taxi'):
        number_taxi = int(value_list[0][4:])
        taxi_coordinateX = int(coordinate_workbook.worksheets[1].cell(row=time + 1, column=number_taxi + 1).value)
        taxi_coordinateY = int(coordinate_workbook.worksheets[2].cell(row=time + 1, column=number_taxi + 1).value)
        firstList.append(taxi_coordinateX)
        firstList.append(taxi_coordinateY)

    elif value_list[0].startswith('bus'):
        number_bus = int(value_list[0][3:])
        bus_coordinateX = int(coordinate_workbook.worksheets[5].cell(row=time + 1, column=number_bus + 1).value)
        bus_coordinateY = int(coordinate_workbook.worksheets[6].cell(row=time + 1, column=number_bus + 1).value)
        firstList.append(bus_coordinateX)
        firstList.append(bus_coordinateY)

    return render_template('index.html', data=data_dict, accuracy=accuracyList, time=time, time_set=time_set, firstList=firstList, secondList=secondList, thirdList=thirdList)

if __name__ == '__main__':
    app.run()